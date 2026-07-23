from pathlib import Path
import sqlite3
import pandas as pd

from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
from statistics import median as calc_median

PROJECT_ROOT = Path(__file__).resolve().parents[2]

DB_PATH = PROJECT_ROOT / "db" / "nifty100.db"

OUTPUT = PROJECT_ROOT / "reports" / "peer_comparison.xlsx"


def load_data():

    conn = sqlite3.connect(DB_PATH)

    df = pd.read_sql("""
    SELECT
        pp.company_id,
        pp.peer_group_name,
        pp.metric,
        pp.percentile_rank,
        pg.is_benchmark
    FROM peer_percentiles pp
    LEFT JOIN peer_groups pg
        ON pp.company_id = pg.company_id
       AND pp.peer_group_name = pg.peer_group_name
    """, conn)

    conn.close()

    return df


def prepare_report(df):

    report = df.pivot_table(
        index=[
            "company_id",
            "peer_group_name"
        ],
        columns="metric",
        values="percentile_rank"
    ).reset_index()

    # Get benchmark information
    benchmark = (
    df.groupby(["company_id", "peer_group_name"])["is_benchmark"]
      .first()
      .reset_index()
      )

   # Merge benchmark column
    report = report.merge(
        benchmark,
        left_on=["company_id", "peer_group_name"],
        right_on=["company_id", "peer_group_name"],
        how="left"
        )

    # Rename columns
    report.rename(columns={
        "company_id": "Company",
        "peer_group_name": "Peer Group",
        "asset_turnover": "Asset Turnover",
        "return_on_equity_pct": "ROE (%)",
        "net_profit_margin_pct": "Net Profit Margin (%)",
        "debt_to_equity": "Debt to Equity",
        "interest_coverage": "Interest Coverage",
        "free_cash_flow_cr": "Free Cash Flow (Cr)",
        "compounded_sales_growth": "Sales CAGR (%)",
        "compounded_profit_growth": "Profit CAGR (%)"
    }, inplace=True)

    return report


def save_excel(report):

    with pd.ExcelWriter(
        OUTPUT,
        engine="openpyxl"
    ) as writer:

        for peer_group in sorted(
            report["Peer Group"].dropna().unique()
        ):

            sheet = report[
                report["Peer Group"] == peer_group
            ]

            sheet.to_excel(
                writer,
                sheet_name=peer_group[:31],
                index=False
            )

    print("Excel report created")
    print(OUTPUT)


def format_workbook():

    from openpyxl import load_workbook

    wb = load_workbook(OUTPUT)

    for ws in wb.worksheets:

        # Header formatting
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                fill_type="solid",
                start_color="1F4E78"
            )
            cell.alignment = Alignment(horizontal="center")

        # Auto-fit columns
        for column in ws.columns:

            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:

                try:
                    if cell.value is not None:
                        max_length = max(
                            max_length,
                            len(str(cell.value))
                        )
                except Exception:
                    pass

            ws.column_dimensions[column_letter].width = max_length + 2

        # Freeze header row
        ws.freeze_panes = "A2"

        # Highlight benchmark company row
        gold_fill = PatternFill(
            fill_type="solid",
            start_color="FFD966"
            )

        for row in range(2, ws.max_row + 1):
            
            benchmark_col = None

            for col in range(1, ws.max_column + 1):
                if ws.cell(row=1, column=col).value == "is_benchmark":
                    benchmark_col = col
                    break

            if benchmark_col is None:
                continue

            if str(ws.cell(row=row, column=benchmark_col).value) == "1":
                
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = gold_fill

        # Hide benchmark column
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == "is_benchmark":
                ws.column_dimensions[get_column_letter(col)].hidden = True


        # Add Median summary row
        median_row = ws.max_row + 1
        
        ws.cell(row=median_row, column=1).value = "Median"
        ws.cell(row=median_row, column=1).font = Font(bold=True)

        for col in range(3, ws.max_column + 1):
            
            values = []

            for row in range(2, median_row):
                
                value = ws.cell(row=row, column=col).value

                if isinstance(value, (int, float)):
                    values.append(value)

            if values:
                median_value = calc_median(values)
                
                cell = ws.cell(row=median_row, column=col)
                cell.value = round(median_value, 2)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(
                    fill_type="solid",
                    start_color="D9EAD3"
                )

        for col in range(1, ws.max_column + 1):
            ws.cell(row=median_row, column=col).font = Font(bold=True)        

        # Apply color scale to percentile columns
        for col in range(3, ws.max_column + 1):

            letter = get_column_letter(col)

            ws.conditional_formatting.add(
                f"{letter}2:{letter}{ws.max_row}",
                ColorScaleRule(
                    start_type="num",
                    start_value=0,
                    start_color="F8696B",   # Red

                    mid_type="num",
                    mid_value=50,
                    mid_color="FFEB84",     # Yellow

                    end_type="num",
                    end_value=100,
                    end_color="63BE7B"      # Green
                )
            )

    wb.save(OUTPUT)


def main():

    df = load_data()

    report = prepare_report(df)

    print(report.head())

    save_excel(report)

    format_workbook()   

    print("Workbook formatting completed.")

if __name__ == "__main__":
    main()